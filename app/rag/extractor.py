from __future__ import annotations
"""
app/rag/extractor.py
Extrai texto puro de arquivos PDF, DOCX, TXT, XLSX e XLS.
"""

import io
from app.core.logging import get_logger

logger = get_logger(__name__)


def extract_text(content: bytes, file_type: str) -> str:
    """
    Ponto de entrada único: detecta o tipo e chama o extrator correto.

    Args:
        content:   bytes do arquivo lido do storage
        file_type: extensão sem ponto em minúsculo — 'pdf', 'docx', 'txt', 'xlsx', 'xls'

    Returns:
        Texto extraído como string; vazio se nada foi encontrado.
    """
    extractors = {
        "pdf":  _extract_pdf,
        "docx": _extract_docx,
        "doc":  _extract_docx,
        "txt":  _extract_txt,
        "xlsx": _extract_excel,
        "xls":  _extract_excel,
    }

    extractor = extractors.get(file_type.lower())
    if not extractor:
        logger.warning("extractor_unsupported_type", file_type=file_type)
        return ""

    try:
        text = extractor(content)
        logger.info("extractor_success", file_type=file_type, chars=len(text))
        return text
    except Exception as exc:
        logger.error("extractor_failed", file_type=file_type, error=str(exc))
        return ""


# ── Extratores individuais ────────────────────────────────────────────────────

def _extract_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages_text = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages_text.append(page_text)
    return "\n\n".join(pages_text)


def _extract_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_txt(content: bytes) -> str:
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _extract_excel(content: bytes) -> str:
    """
    Extrai texto de planilhas Excel (.xlsx e .xls) usando openpyxl.

    Cada aba é processada separadamente. As linhas são convertidas em texto
    no formato "Coluna1: Valor1 | Coluna2: Valor2 ..." usando a primeira linha
    como cabeçalho. Isso preserva o contexto das colunas para o RAG entender
    o significado de cada valor.

    Exemplo de saída para uma linha:
        Produto: Caneta | Quantidade: 100 | Preço: 2.50 | Total: 250.00
    """
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    all_text = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows  = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        # Primeira linha como cabeçalho — remove células vazias
        headers = [str(h).strip() if h is not None else f"Coluna{i+1}"
                   for i, h in enumerate(rows[0])]

        sheet_lines = [f"[Aba: {sheet_name}]"]

        for row in rows[1:]:
            # Ignora linhas completamente vazias
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # Monta "Cabeçalho: Valor" para cada célula não vazia
            parts = []
            for header, cell in zip(headers, row):
                if cell is not None and str(cell).strip() != "":
                    parts.append(f"{header}: {cell}")

            if parts:
                sheet_lines.append(" | ".join(parts))

        if len(sheet_lines) > 1:   # só adiciona se tiver dados além do cabeçalho
            all_text.append("\n".join(sheet_lines))

    return "\n\n".join(all_text)
